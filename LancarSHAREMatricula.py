import os
import time
import json
import pyautogui
import pyperclip
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# CONFIGURAÇÕES

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PASTA_SAIDA = os.path.join(BASE_DIR, "saida_excel")
PASTA_CONFIG = os.path.join(BASE_DIR, "config")
ARQUIVO_COORDENADAS = os.path.join(PASTA_CONFIG, "coordenadas.json")

LINHA_INICIAL = 2

TEMPO_ENTRE_ACOES = 0.6
TEMPO_APOS_CONFIRMAR = 1.3

pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.2

# CORES EXCEL

VERDE_OK = PatternFill(fill_type="solid", start_color="C6E0B4", end_color="C6E0B4")
VERMELHO_ERRO = PatternFill(fill_type="solid", start_color="F4CCCC", end_color="F4CCCC")

# FUNÇÕES

def carregar_coordenadas():
    if not os.path.exists(ARQUIVO_COORDENADAS):
        print("Arquivo de coordenadas não encontrado.")
        print("Execute primeiro: python AtualizarCoordenadas.py")
        input("\nPressione Enter para sair...")
        raise SystemExit

    with open(ARQUIVO_COORDENADAS, "r", encoding="utf-8") as f:
        dados = json.load(f)

    return {
        "valor_estimado": (dados["valor_estimado"]["x"], dados["valor_estimado"]["y"]),
        "obs": (dados["obs"]["x"], dados["obs"]["y"]),
        "pasta_processo": (dados["pasta_processo"]["x"], dados["pasta_processo"]["y"]),
        "confirmar": (dados["confirmar"]["x"], dados["confirmar"]["y"]),
        "docto": (dados["docto"]["x"], dados["docto"]["y"]),
    }

coords = carregar_coordenadas()

POS_VALOR_ESTIMADO = coords["valor_estimado"]
POS_OBS = coords["obs"]
POS_PASTA = coords["pasta_processo"]
POS_CONFIRMAR = coords["confirmar"]
POS_DOCTO = coords["docto"]

def colar_texto(texto):
    pyperclip.copy(str(texto))
    pyautogui.hotkey("ctrl", "v")
    time.sleep(TEMPO_ENTRE_ACOES)

def clicar_e_colar(posicao, texto):
    pyautogui.click(posicao)
    time.sleep(0.2)
    colar_texto(texto)

def encontrar_planilha_mais_recente():
    arquivos = [
        f for f in os.listdir(PASTA_SAIDA)
        if f.lower().endswith(".xlsx") and f.startswith("Reembolso_ARISP_")
    ]

    if not arquivos:
        return None

    arquivos.sort(reverse=True)
    return os.path.join(PASTA_SAIDA, arquivos[0])

# LOCALIZAR PLANILHA

CAMINHO_PLANILHA = encontrar_planilha_mais_recente()

if not CAMINHO_PLANILHA or not os.path.exists(CAMINHO_PLANILHA):
    print("Nenhuma planilha válida encontrada em 'saida_excel'.")
    input("\nPressione Enter para sair...")
    raise SystemExit

# LER PLANILHA

df = pd.read_excel(CAMINHO_PLANILHA, dtype=str).fillna("")

wb = load_workbook(CAMINHO_PLANILHA)
ws = wb.active

print("\n======================================")
print(" LANÇADOR SHARE - REEMBOLSOS ARISP ")
print("======================================\n")

print(f"Planilha usada:\n{os.path.basename(CAMINHO_PLANILHA)}\n")

ja_feitas = 0
pendentes = 0

for idx in range(LINHA_INICIAL - 2, len(df)):
    status_extracao = str(df.loc[idx, "Status Extração"]).strip().upper()
    status_share = str(df.loc[idx, "Status SHARE"]).strip().upper()

    if status_share == "OK":
        ja_feitas += 1
    elif status_extracao == "OK" and status_share != "OK":
        pendentes += 1

print(f"Linhas já lançadas no SHARE: {ja_feitas}")
print(f"Linhas pendentes para lançar: {pendentes}\n")

print("Você tem 10 segundos para deixar a tela do SHARE pronta.")
print("Abra a tela de lançamento e NÃO mexa no mouse/teclado.")
print("Se precisar parar, jogue o mouse em um canto da tela.\n")
time.sleep(10)

primeira_linha_executada = True

try:
    for idx in range(LINHA_INICIAL - 2, len(df)):
        linha_excel = idx + 2

        tipo = str(df.loc[idx, "Tipo"]).strip()
        matricula = str(df.loc[idx, "Matricula"]).strip()
        pasta = str(df.loc[idx, "Pasta"]).strip()
        valor_total = str(df.loc[idx, "Valor Total"]).strip()
        status_extracao = str(df.loc[idx, "Status Extração"]).strip().upper()
        status_share = str(df.loc[idx, "Status SHARE"]).strip().upper()

        if pasta.endswith(".0"):
            pasta = pasta[:-2]

        if status_share == "OK":
            print(f"[Linha {linha_excel}] Já lançada anteriormente. Pulando...")
            continue

        if status_extracao != "OK":
            print(f"[Linha {linha_excel}] Extração com erro. Pulando...")
            continue

        print(f"\n[Linha {linha_excel}] PROCESSANDO")
        print(f"Tipo = {tipo}")
        print(f"Matrícula = {matricula}")
        print(f"Pasta = {pasta}")
        print(f"Valor Total = {valor_total}")

        if not primeira_linha_executada:
            pyautogui.click(POS_DOCTO)
            time.sleep(0.8)

        clicar_e_colar(POS_VALOR_ESTIMADO, valor_total)
        clicar_e_colar(POS_OBS, matricula)
        clicar_e_colar(POS_PASTA, pasta)

        pyautogui.click(POS_CONFIRMAR)
        time.sleep(TEMPO_APOS_CONFIRMAR)

        ws[f"G{linha_excel}"] = "OK"

        for col in range(1, 8):
            ws.cell(row=linha_excel, column=col).fill = VERDE_OK

        wb.save(CAMINHO_PLANILHA)

        print("Confirmado. Linha salva como OK no SHARE.")

        primeira_linha_executada = False

except pyautogui.FailSafeException:
    print("\n\n======================================")
    print(" PARADA MANUAL DETECTADA ")
    print("======================================")
    print("Você moveu o mouse para um canto da tela.")
    print("O processo foi interrompido com segurança.")
    print("Tudo que já estava salvo continuará salvo.")
    wb.save(CAMINHO_PLANILHA)
    input("\nPressione Enter para sair...")

except Exception as e:
    print(f"\nERRO inesperado: {e}")
    print("O processo foi interrompido para evitar problema maior.")
    wb.save(CAMINHO_PLANILHA)
    input("\nPressione Enter para sair...")

wb.save(CAMINHO_PLANILHA)

print("\n======================================")
print(" PROCESSO FINALIZADO ")
print("======================================")
print("Todas as linhas pendentes possíveis foram processadas.")
print("A planilha foi atualizada.")
input("\nPressione Enter para sair...")